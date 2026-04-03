import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Top-level configuration
USERS_URL = "https://jsonplaceholder.typicode.com/users"
OUTPUT_XLSX = "users.xlsx"
TIMEOUT = 10
HEADERS = ["id", "name", "email", "phone", "website", "company_name"]


def fetch_users():
    """
    Fetch users from the API.

    Returns a minimal result dict:
    {
        "ok": bool,
        "stage": "fetch_users",
        "reason": str,
        "evidence": any (supporting info only),
        "data": any (raw data to pass forward, when ok=True)
    }
    """
    result = {
        "ok": False,
        "stage": "fetch_users",
        "reason": "",
        "evidence": None,
        "data": None,
    }

    try:
        response = requests.get(USERS_URL, timeout=TIMEOUT)
        response.raise_for_status()
    except requests.RequestException as exc:
        result["reason"] = f"Request failed: {exc}"
        return result

    try:
        data = response.json()
    except ValueError as exc:
        result["reason"] = f"Failed to parse JSON: {exc}"
        return result

    if not isinstance(data, list):
        result["reason"] = "Unexpected response format (expected a list)."
        result["evidence"] = {"actual_type": type(data).__name__}
        return result

    result["ok"] = True
    result["reason"] = "Users fetched successfully."
    result["data"] = data  # raw API data only in 'data'
    result["evidence"] = {"count": len(data)}
    return result


def extract_rows(users):
    """
    Transform raw user JSON objects into flat row dicts for Excel.

    Returns a result dict:
    {
        "ok": bool,
        "stage": "extract_rows",
        "reason": str,
        "evidence": any (supporting info only),
        "data": list[dict] (rows ready for Excel, when ok=True)
    }
    """
    result = {
        "ok": False,
        "stage": "extract_rows",
        "reason": "",
        "evidence": None,
        "data": None,
    }

    rows = []
    skipped = 0

    for index, user in enumerate(users, start=1):
        try:
            row = {
                "id": user["id"],
                "name": user["name"],
                "email": user["email"],
                "phone": user["phone"],
                "website": user["website"],
                "company_name": user["company"]["name"],
            }
            rows.append(row)
        except (KeyError, TypeError):
            skipped += 1

    if not rows:
        result["reason"] = "No valid user records to extract."
        result["evidence"] = {"skipped": skipped}
        return result

    result["ok"] = True
    result["reason"] = "Rows extracted successfully."
    result["data"] = rows
    result["evidence"] = {
        "rows_count": len(rows),
        "skipped": skipped,
    }
    return result


def save_rows_to_excel(rows):
    """
    Save already-prepared rows to an Excel file.

    Returns a result dict:
    {
        "ok": bool,
        "stage": "save_rows_to_excel",
        "reason": str,
        "evidence": any (summary info only),
        "data": None
    }
    """
    result = {
        "ok": False,
        "stage": "save_rows_to_excel",
        "reason": "",
        "evidence": None,
        "data": None,
    }

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "users"

        ws.append(HEADERS)

        for row in rows:
            ws.append([row[h] for h in HEADERS])

        # Enable AutoFilter on header row
        last_col = get_column_letter(len(HEADERS))
        last_row = len(rows) + 1
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

        wb.save(OUTPUT_XLSX)
    except Exception as exc:
        result["reason"] = f"Failed to write Excel file: {exc}"
        result["evidence"] = {"rows_attempted": len(rows)}
        return result

    result["ok"] = True
    result["reason"] = "Excel file written successfully."
    result["evidence"] = {
        "rows_written": len(rows),
        "output_file": OUTPUT_XLSX,
    }
    return result


def main():
    print("Starting user export to Excel...\n")

    # Step 1: fetch
    fetch_result = fetch_users()
    print(
        f"[{fetch_result['stage']}] ok={fetch_result['ok']} "
        f"reason='{fetch_result['reason']}' "
        f"evidence={fetch_result['evidence']}"
    )
    if not fetch_result["ok"]:
        print("Stopping because fetching users failed.")
        return 1

    users = fetch_result["data"]

    # Step 2: extract
    extract_result = extract_rows(users)
    print(
        f"[{extract_result['stage']}] ok={extract_result['ok']} "
        f"reason='{extract_result['reason']}' "
        f"evidence={extract_result['evidence']}"
    )
    if not extract_result["ok"]:
        print("Stopping because row extraction failed.")
        return 1

    rows = extract_result["data"]

    # Step 3: save
    save_result = save_rows_to_excel(rows)
    print(
        f"[{save_result['stage']}] ok={save_result['ok']} "
        f"reason='{save_result['reason']}' "
        f"evidence={save_result['evidence']}"
    )
    if not save_result["ok"]:
        print("Saving rows to Excel failed.")
        return 1

    print("\nUser export to Excel finished successfully.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())